import pandas as pd
import pdfplumber
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def extract_pdf_data(pdf_path):
    """从PDF中提取国家信息"""
    country_data = {}
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                
                for row in table:
                    if not row or len(row) < 5:
                        continue
                    
                    country = row[0]
                    if not country or country in ['Country', 'Service', '']:
                        continue
                    
                    country = country.strip()
                    service_type = row[1].strip() if len(row) > 1 and row[1] else ''
                    number_avail = row[2].strip() if len(row) > 2 and row[2] else ''
                    national_out = row[3].strip() if len(row) > 3 and row[3] else ''
                    intl_out = row[4].strip() if len(row) > 4 and row[4] else ''
                    porting = row[5].strip() if len(row) > 5 and row[5] else ''
                    
                    key = (normalize_country_name(country), normalize_service_type(service_type))
                    if key not in country_data:
                        country_data[key] = {
                            'Service Type': service_type,
                            'Number availability': number_avail,
                            'National outbound': national_out,
                            'international outbound': intl_out,
                            'Porting Available': porting
                        }
    
    return country_data

def normalize_country_name(name):
    """标准化国家名称"""
    if not name or pd.isna(name):
        return ''
    name = str(name).strip().lower()
    name = re.sub(r'\s+', ' ', name)
    
    mappings = {
        'us': 'united states',
        'usa': 'united states',
        'uk': 'united kingdom',
        'hong kong & macao': 'hong kong',
        'korea': 'south korea'
    }
    
    return mappings.get(name, name)

def normalize_service_type(service_type):
    """标准化服务类型"""
    if not service_type or pd.isna(service_type):
        return ''
    st = str(service_type).strip().lower()
    if 'toll' in st and 'free' in st:
        return 'toll-free'
    elif 'did' in st:
        return 'did'
    return st

def match_country_and_type(excel_country, excel_type, pdf_data):
    """匹配Excel中的国家名和服务类型与PDF数据"""
    excel_country_norm = normalize_country_name(excel_country)
    excel_type_norm = normalize_service_type(excel_type)
    
    for (pdf_country_norm, pdf_type_norm), data in pdf_data.items():
        if excel_country_norm == pdf_country_norm and excel_type_norm == pdf_type_norm:
            return data
        if excel_country_norm in pdf_country_norm or pdf_country_norm in excel_country_norm:
            if excel_type_norm == pdf_type_norm:
                return data
    
    return None

def update_excel(excel_path, pdf_path):
    """更新Excel文件"""
    print("正在读取PDF文件...")
    pdf_data = extract_pdf_data(pdf_path)
    print(f"从PDF中提取了 {len(pdf_data)} 个国家的数据")
    
    print("\n正在读取Excel文件...")
    df = pd.read_excel(excel_path, sheet_name='Phone Number and Type', header=None)
    
    header_row = 1
    df.columns = df.iloc[header_row]
    df = df.iloc[header_row+1:].reset_index(drop=True)
    
    country_col = df.columns[0]
    
    col_mapping = {}
    service_type_col = None
    for col in df.columns:
        col_str = str(col).strip()
        if col_str == 'Number availability':
            col_mapping['Number availability'] = col
        elif col_str == 'National outbound':
            col_mapping['National outbound'] = col
        elif col_str == 'international outbound':
            col_mapping['international outbound'] = col
        elif 'Porting' in col_str:
            col_mapping['Porting Available'] = col
        elif 'Numbery Type' in col_str or 'Number Type' in col_str:
            service_type_col = col
    
    print(f"\n找到的列映射: {col_mapping}")
    print(f"服务类型列: {service_type_col}")
    
    updates = 0
    updated_cells = []
    for idx, row in df.iterrows():
        country = row[country_col]
        if pd.isna(country) or country == '':
            continue
        
        service_type = row[service_type_col] if service_type_col and service_type_col in df.columns else ''
        
        country_info = match_country_and_type(country, service_type, pdf_data)
        
        if country_info:
            for pdf_key, excel_col in col_mapping.items():
                if excel_col in df.columns and pdf_key in country_info:
                    old_val = str(row[excel_col]).strip() if not pd.isna(row[excel_col]) else ''
                    new_val = str(country_info[pdf_key]).strip()
                    
                    if old_val != new_val and new_val:
                        df.at[idx, excel_col] = new_val
                        col_idx = df.columns.get_loc(excel_col)
                        updated_cells.append((idx + header_row + 2, col_idx + 1))
                        updates += 1
                        print(f"更新 {country} ({service_type}) 的 {pdf_key}: '{old_val}' -> '{new_val}'")
    
    print(f"\n共更新了 {updates} 个单元格")
    
    original_df = pd.read_excel(excel_path, sheet_name='Phone Number and Type', header=None)
    original_df.iloc[header_row+1:] = df.values
    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        original_df.to_excel(writer, sheet_name='Phone Number and Type', index=False, header=False)
    
    wb = load_workbook(excel_path)
    ws = wb['Phone Number and Type']
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    for row_idx, col_idx in updated_cells:
        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
    
    wb.save(excel_path)
    print(f"\n已保存更新到 {excel_path}，更新的单元格已高亮显示")

if __name__ == '__main__':
    excel_file = 'phone_number.xlsx'
    pdf_file = 'Amazon_Connect_Telecoms_Coverage.pdf'
    
    update_excel(excel_file, pdf_file)
